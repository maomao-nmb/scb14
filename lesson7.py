'''接口自动化步骤
1、excel测试用例准备ok，代码自动读取测试数据  ---read_data()
2、发送接口请求，得到响应信息  ---api_fun()
3、断言：实际结果 vs预期结果   ---通过/不通过
4、写入excel，通过/不通过

函数：eval()  ---运行被字符串包裹的表达式
 '{"mobile_phone":"13652440101","pwd":"lemon666","type":0,"reg_name":"LemonScb14"}'


'''



import requests
import openpyxl

#读取测试用例函数
def read_data(filename,sheetname):
    wb = openpyxl.load_workbook(filename)
    sheet = wb[sheetname]
    max_row =sheet.max_row  #获取最大行数
    case_list =[]  #创建空列表，存放测试用例
    for i in range(2,max_row+1):
        dict1=dict(
        case_id= sheet.cell(row=i, column=1).value,   #获取表单case_id的值
        url = sheet.cell(row=i, column=5).value,      #获取表单url的值
        data = sheet.cell(row=i, column=6).value,
        expected = sheet.cell(row=i, column=7).value,
        )
        case_list.append(dict1) #每循环一次，就把读取到的字典数据存放到这个list
    return case_list   #返回测试用例列表 --得到数据，函数结束

#执行接口函数
def api_fun(url,data):
    headers_login = {'X-Lemonban-Media-Type':'lemonban.v2','Content-Type':'application/json'}   # 请求头，字典的形式
    res = requests.post(url=url, json=data, headers=headers_login)  #接收post方法请求的结果
    response = res.json()   # 响应的正文
    return response

#写入结果函数
def write_result(filename,sheetname,row,column,final_result):
    wb = openpyxl.load_workbook(filename)
    sheet = wb[sheetname]
    sheet.cell(row=row, column=column).value =final_result   # 直接写入结果
    wb.save(filename)  # 保存，关


# cases =read_data('test_case_api.xlsx','register')   #调用读取测试用例函数，获取所有的测试用例数据保存到变量cases
# for case in cases:
#     case_id = case.get('case_id') # 也可以通过case['case_id']取
#     url = case.get('url')
#     data = eval(case.get('data'))  # eval() 运行被字符串包裹的表达式 ---去掉字符串引号
#     expected = eval(case.get('expected')) # 获取预期结果
#     expected_msg = expected.get('msg') # 获取预期结果中的msg
#     # print(data)
#     # print(type(data))
#     # print(case_id,url,data,expected)
#     real_result= api_fun(url=url, data=data) #调用执行接口函数（发送接口请求函数），返回结果用变量real_result接收
#     real_msg = real_result.get('msg') # 获取实际结果中的msg
#     print('预期结果中的msg:{}'.format(expected_msg))
#     print('实际结果中的msg:{}'.format(real_msg))
#     if real_msg == expected_msg:
#         print('第{}条测试用例执行通过!'.format(case_id))
#         final_re = 'passed'
#     else:
#         print('第{}条测试用例执行不通过!'.format(case_id))
#         final_re = 'failed'
#     write_result('test_case_api.xlsx','register',case_id+1 , 8, final_re)
#     print('*'*20)
#



# 执行测试用例并回写实际结果函数
def execute_fun(filename,sheetname):
    cases =read_data(filename,sheetname)   #调用读取测试用例函数，获取所有的测试用例数据保存到变量cases
    for case in cases:
        case_id = case.get('case_id') # 也可以通过case['case_id']取
        url = case.get('url')
        data = eval(case.get('data'))  # eval() 运行被字符串包裹的表达式 ---去掉字符串引号
        expected = eval(case.get('expected')) # 获取预期结果
        expected_msg = expected.get('msg') # 获取预期结果中的msg
        # print(data)
        # print(type(data))
        # print(case_id,url,data,expected)
        real_result= api_fun(url=url, data=data) #调用执行接口函数（发送接口请求函数），返回结果用变量real_result接收
        real_msg = real_result.get('msg') # 获取实际结果中的msg
        print('预期结果中的msg:{}'.format(expected_msg))
        print('实际结果中的msg:{}'.format(real_msg))
        if real_msg == expected_msg:
            print('第{}条测试用例执行通过!'.format(case_id))
            final_re = 'passed'
        else:
            print('第{}条测试用例执行不通过!'.format(case_id))
            final_re = 'failed'
        write_result(filename,sheetname,case_id+1 , 8, final_re)
        print('*'*20)

execute_fun('test_case_api.xlsx','login')  #调用执行测试用例并回写实际结果的函数




# # 函数：eval()  ---运行被字符串包裹的表达式
# a = '{"mobile_phone":"13652440101","pwd":"lemon666","type":0,"reg_name":"LemonScb14"}'
# b = eval(a)
# print(b)



